import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

TEAL = "2A6B6E"
TEAL_TINT = "E3EEEE"
INK = "1C1C1C"
MUTED = "6B6B6B"
YELLOW = "FFF6D6"

FONT_NAME = "Calibri"
header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
title_font = Font(name=FONT_NAME, size=18, bold=True, color=TEAL)
sub_font = Font(name=FONT_NAME, size=10.5, color=MUTED, italic=True)
body_font = Font(name=FONT_NAME, size=10.5, color=INK)
example_font = Font(name=FONT_NAME, size=10.5, color=MUTED, italic=True)
label_font = Font(name=FONT_NAME, size=10.5, bold=True, color=INK)

header_fill = PatternFill("solid", fgColor=TEAL)
tint_fill = PatternFill("solid", fgColor=TEAL_TINT)
yellow_fill = PatternFill("solid", fgColor=YELLOW)

thin = Side(style="thin", color="D9D6CC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_OPTIONS = ["Not Yet Applied", "Applied", "Phone Screen", "Interview Scheduled",
                   "Interviewed", "Offer", "Rejected", "Withdrawn"]

# ---------------------------------------------------------------- Tracker sheet
ws = wb.active
ws.title = "Application Tracker"
ws.sheet_view.showGridLines = False

headers = ["Company", "Role / Title", "Date Applied", "Source", "Status",
           "Contact Person", "Contact Email", "Follow-Up Date", "Salary Range", "Notes"]
widths = [22, 26, 14, 16, 18, 18, 24, 14, 16, 34]

ws.merge_cells("A1:J1")
ws["A1"] = "Job Application Tracker"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:J2")
ws["A2"] = "Log every application here. Fill in the yellow example row's format, then delete it before you start."
ws["A2"].font = sub_font
ws.row_dimensions[2].height = 18

header_row = 4
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
ws.row_dimensions[header_row].height = 30

example = ["Meridian Health Group", "Clinical Operations Manager", "2026-01-15", "LinkedIn",
           "Interview Scheduled", "Priya Anand", "panand@meridianhealth.com", "2026-01-25",
           "$85,000–$95,000", "Referred by former colleague; second interview is a panel."]
for i, val in enumerate(example, start=1):
    c = ws.cell(row=5, column=i, value=val)
    c.font = example_font
    c.fill = yellow_fill
    c.border = border
    c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[5].height = 30

# 60 blank rows ready to fill in
for r in range(6, 66):
    for i in range(1, 11):
        c = ws.cell(row=r, column=i)
        c.font = body_font
        c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 20

ws.freeze_panes = "A5"

dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTIONS) + '"', allow_blank=True)
ws.add_data_validation(dv_status)
dv_status.add(f"E5:E200")

dv_date = DataValidation(type="date", operator="greaterThan", formula1="1900-01-01",
                          allow_blank=True, showErrorMessage=True,
                          errorTitle="Invalid date", error="Please enter a valid date.")
ws.add_data_validation(dv_date)
dv_date.add("C5:C200")
dv_date.add("H5:H200")

# ---------------------------------------------------------------- Dashboard sheet
ws2 = wb.create_sheet("Dashboard")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:D1")
ws2["A1"] = "Search Overview"
ws2["A1"].font = title_font
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:D2")
ws2["A2"] = "Updates automatically from the Application Tracker sheet."
ws2["A2"].font = sub_font

labels = ["Total Applications", "Awaiting Response", "In Interview Process", "Offers", "Rejected / Closed"]
formulas = [
    '=COUNTA(\'Application Tracker\'!A6:A200)',
    '=COUNTIF(\'Application Tracker\'!E6:E200,"Applied")',
    '=COUNTIF(\'Application Tracker\'!E6:E200,"Phone Screen")+COUNTIF(\'Application Tracker\'!E6:E200,"Interview Scheduled")+COUNTIF(\'Application Tracker\'!E6:E200,"Interviewed")',
    '=COUNTIF(\'Application Tracker\'!E6:E200,"Offer")',
    '=COUNTIF(\'Application Tracker\'!E6:E200,"Rejected")+COUNTIF(\'Application Tracker\'!E6:E200,"Withdrawn")',
]
row = 4
for label, formula in zip(labels, formulas):
    ws2.cell(row=row, column=1, value=label).font = label_font
    ws2.cell(row=row, column=1).fill = tint_fill
    ws2.cell(row=row, column=1).border = border
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    v = ws2.cell(row=row, column=4, value=formula)
    v.font = Font(name=FONT_NAME, size=12, bold=True, color=TEAL)
    v.alignment = Alignment(horizontal="center")
    v.border = border
    ws2.row_dimensions[row].height = 22
    row += 1

for col, w in zip("ABCD", [22, 10, 10, 12]):
    ws2.column_dimensions[col].width = w

# ---------------------------------------------------------------- Instructions sheet
ws3 = wb.create_sheet("Instructions")
ws3.sheet_view.showGridLines = False
ws3.merge_cells("A1:B1")
ws3["A1"] = "How to Use This Tracker"
ws3["A1"].font = title_font
ws3.row_dimensions[1].height = 28
ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 90

lines = [
    "1. Go to the \"Application Tracker\" tab. The first data row (highlighted yellow) shows an example — copy its format, then delete it.",
    "2. Add one row per application. The Status column is a dropdown — click any cell in that column to pick from the list.",
    "3. Use the Follow-Up Date column to remind yourself when to check back if you haven't heard anything (7–10 days after applying is typical).",
    "4. The \"Dashboard\" tab totals your applications by stage automatically — nothing to update by hand.",
    "5. Sort or filter the table (Data > Filter in Excel/Sheets) to see everything at a given status, e.g. every application still \"Awaiting Response.\"",
    "",
    "This tracker pairs with the Job Search Playbook guide included in this bundle — see that guide for the full search system this sheet is built around.",
]
r = 3
for line in lines:
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    c = ws3.cell(row=r, column=2, value=line)
    c.font = body_font
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[r].height = 34 if line else 10
    r += 1

wb.move_sheet("Instructions", offset=-2)

wb.save("/home/user/saimakhalid/product/04-Job-Application-Tracker.xlsx")
print("saved")
